#!/usr/bin/env python3
"""Gera os JSONs estáticos consumidos pelo dashboard no GitHub Pages."""

from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT / "dados" / "17. Atualização de SCAs"
OUTPUT_ROOT = ROOT / "dados-processados"
CONTRACT = json.loads((ROOT / "config" / "dashboard_contract.json").read_text(encoding="utf-8"))
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
UNITS = {
    "nova-mutum": {"folder": "1. Nova Mutum", "label": "Nova Mutum / MT"},
    "rondonopolis": {"folder": "2. Rondonópolis", "label": "Rondonópolis / MT"},
    "rio-verde": {"folder": "3. Rio Verde", "label": "Rio Verde / GO"},
}
WEIGHT = {"Concluído": 1.0, "Em andamento": 0.5, "Não iniciado": 0.0}


def clean(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    return re.sub(r"\s+", " ", str(value)).strip() or default


def clean_cell(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def person_name(value: Any, default: str) -> str:
    text = clean(value, default).title()
    for particle in (" Da ", " De ", " Do ", " Das ", " Dos "):
        text = text.replace(particle, particle.lower())
    return text


def key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value).casefold())
    return re.sub(r"[^a-z0-9]+", " ", "".join(c for c in text if not unicodedata.combining(c))).strip()


def fmt_date(value: Any) -> str | None:
    if value in (None, "", 0, "00:00:00"):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    text = clean(value)
    for pattern in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return text


def iso_date(value: Any) -> str | None:
    if value in (None, "", 0, "00:00:00"):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    for pattern in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def valid_iso(value: str | None) -> bool:
    return bool(value and re.fullmatch(r"\d{4}-\d{2}-\d{2}", value))


def item_date(value: Any, item_code: str, field: str, contract: dict[str, Any]) -> str | None:
    parsed = iso_date(value)
    if valid_iso(parsed):
        return parsed
    return contract.get("item_dates", {}).get(item_code, {}).get(field)


def normalize_status(value: Any) -> str:
    token = key(value)
    if token in {"concluido", "concluida", "3"}:
        return "Concluído"
    if token in {"em andamento", "andamento", "2"}:
        return "Em andamento"
    if token in {"nao iniciado", "nao iniciada", "1"}:
        return "Não iniciado"
    return "Não iniciado"


def workbook_files(folder: Path) -> list[Path]:
    return sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~$") and p.is_file())


def first_workbook(folder: Path) -> Path | None:
    files = workbook_files(folder)
    return files[0] if files else None


def read_sheet(path: Path, preferred: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    return wb[wb.sheetnames[0]]


def find_header(ws, required: set[str], max_rows: int = 30) -> tuple[int, list[str]] | None:
    wanted = {key(x) for x in required}
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row), values_only=True), 1):
        headers = [clean(x) for x in row]
        if wanted.issubset({key(x) for x in headers if x}):
            return row_number, headers
    return None


def header_index(headers: list[str], *names: str) -> int | None:
    lookup = {key(v): i for i, v in enumerate(headers) if v}
    for name in names:
        if key(name) in lookup:
            return lookup[key(name)]
    return None


def value_at(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def classify_image(name: str) -> str:
    token = key(Path(name).stem)
    if re.search(r"(^| )projeto( |$)", token):
        return "Projeto"
    if re.search(r"(^| )(obra|campo)( |$)", token):
        return "Obra"
    return "Obra"


def build_images(unit_id: str, unit_folder: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, list[str]]]]:
    image_root = unit_folder / "1. Imagens"
    manifest: list[dict[str, Any]] = []
    by_area: dict[str, dict[str, list[str]]] = defaultdict(lambda: {"Projeto": [], "Obra": []})
    if not image_root.exists():
        return manifest, by_area
    for path in sorted(image_root.rglob("*"), key=lambda p: key(p.as_posix())):
        if not path.is_file() or path.suffix.casefold() not in IMAGE_EXTENSIONS:
            continue
        relative_area = path.parent.relative_to(image_root).as_posix()
        area = relative_area.split("/")[0] if relative_area != "." else "Sem área"
        kind = classify_image(path.name)
        repo_path = path.relative_to(ROOT).as_posix()
        url = "/".join(quote(part, safe="-._~()'") for part in repo_path.split("/"))
        item = {"unidade": unit_id, "area": area, "arquivo": path.name, "caminho": url, "classificacao": kind}
        manifest.append(item)
        by_area[key(area)][kind].append(url)
    return manifest, by_area


def aggregate_status(rows: list[dict[str, Any]]) -> tuple[int, int, int, int, str]:
    counts = Counter(r["situacao"] for r in rows)
    total = len(rows)
    progress = round(100 * sum(WEIGHT.get(r["situacao"], 0) for r in rows) / total) if total else 0
    if total and counts["Concluído"] == total:
        status = "Concluído"
    elif total and counts["Não iniciado"] == total:
        status = "Não iniciado"
    else:
        status = "Em andamento" if total else "Não iniciado"
    return counts["Concluído"], counts["Em andamento"], counts["Não iniciado"], progress, status


def latest_date(values: list[Any]) -> str | None:
    dates = [fmt_date(v) for v in values]
    dates = [d for d in dates if d]
    if not dates:
        return None
    def parse(v: str):
        try:
            return datetime.strptime(v, "%d/%m/%Y")
        except ValueError:
            return datetime.min
    return max(dates, key=parse)


def empty_operational(label: str) -> dict[str, Any]:
    return {
        "meta": {"unidade": label, "fases": "—", "conteudo": "Sistemas de Controle Ambiental (SCA)", "responsavel": "Não informado", "atualizacao": "Sem dados disponíveis", "prazo_protocolo": "—", "prazo_docs": "—", "vistoria": "—", "termino": "—"},
        "totais": {"total_areas": 0, "ilhas": 0, "total_scas": 0, "areas_andamento": 0, "areas_nao_iniciadas": 0, "areas_concluidas": 0, "itens_concluidos": 0, "itens_andamento": 0, "itens_nao_iniciados": 0},
        "fase": {}, "reprog": {"scas": 0, "pct": 0, "total": 0}, "macro": [], "disciplina": [], "sca_macro": [], "gerente": [], "areas": []
    }


def read_operational(path: Path | None, label: str, image_map: dict[str, dict[str, list[str]]], contract: dict[str, Any]) -> dict[str, Any]:
    result = empty_operational(label)
    if not path:
        return result
    ws = read_sheet(path, "SCAs")
    result["meta"].update({
        "unidade": label,
        "fases": clean(ws["E3"].value, "Não informado"),
        "conteudo": clean(ws["E4"].value, "Sistemas de Controle Ambiental (SCA)"),
        "responsavel": clean(ws["E6"].value, "Não informado").title(),
        "atualizacao": fmt_date(ws["E7"].value) or "Não informado",
    })
    found = find_header(ws, {"Área Descrição MAM", "Sistema de Controle Ambiental", "Situação"})
    if not found:
        return result
    header_row, headers = found
    indexes = {
        "proc_cod": header_index(headers, "Cód. do Processo - Engenharia"),
        "macro": header_index(headers, "Macro área"), "ilha": header_index(headers, "Ilha de Processo"),
        "processo": header_index(headers, "Processo"), "area": header_index(headers, "Área Descrição MAM"),
        "fase": header_index(headers, "Fase"), "vistoria": header_index(headers, "Prazo p/ vistoria"),
        "previsto": header_index(headers, "Previsto (LB)"), "replan": header_index(headers, "Replan"),
        "replans": header_index(headers, "Qtde de replans"), "item_cod": header_index(headers, "Cód. EAP"),
        "item": header_index(headers, "Sistema de Controle Ambiental"), "sca_macro": header_index(headers, "SCA Macro"),
        "disciplina": header_index(headers, "Disciplina"), "supervisor": header_index(headers, "Supervisor responsável"),
        "gerente": header_index(headers, "Gerente responsável"), "situacao": header_index(headers, "Situação"),
        "observacao": header_index(headers, "Observação"),
    }
    raw: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_name = clean(value_at(row, indexes["item"]))
        area_name = clean(value_at(row, indexes["area"]))
        if not item_name or not area_name:
            continue
        raw.append({
            "proc_cod": clean(value_at(row, indexes["proc_cod"])), "macro": clean(value_at(row, indexes["macro"]), "Não informado"),
            "ilha": clean(value_at(row, indexes["ilha"]), "Não informado"), "processo": clean(value_at(row, indexes["processo"]), area_name),
            "area": area_name, "fase": clean(value_at(row, indexes["fase"]), "Não informado"),
            "vistoria": value_at(row, indexes["vistoria"]), "previsto": value_at(row, indexes["previsto"]), "replan": value_at(row, indexes["replan"]),
            "replans": value_at(row, indexes["replans"]), "item_cod": clean(value_at(row, indexes["item_cod"])), "item": item_name,
            "sca_macro": clean(value_at(row, indexes["sca_macro"]), "Não informado"), "disciplina": clean(value_at(row, indexes["disciplina"]), "Não informado"),
            "supervisor": clean(value_at(row, indexes["supervisor"]), "Não informado"), "gerente": clean(value_at(row, indexes["gerente"]), "Não informado"),
            "situacao": normalize_status(value_at(row, indexes["situacao"])), "observacao": clean(value_at(row, indexes["observacao"])),
        })
    if not raw:
        return empty_operational(label)
    result["meta"]["vistoria"] = latest_date([r["vistoria"] for r in raw]) or "Não informado"
    result["meta"]["termino"] = latest_date([r["replan"] or r["previsto"] for r in raw]) or "Não informado"
    result["meta"]["prazo_protocolo"] = result["meta"]["termino"]
    result["meta"]["prazo_docs"] = result["meta"]["termino"]

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    area_names: dict[str, str] = {}
    for item in raw:
        area_names.setdefault(key(item["area"]), item["area"])
        grouped[key(item["area"])].append(item)
    areas: list[dict[str, Any]] = []
    for number, area_key in enumerate(grouped, 1):
        rows = grouped[area_key]
        conc, andamento, nao, progresso, status = aggregate_status(rows)
        photos = image_map.get(area_key, {"Projeto": [], "Obra": []})
        item_macro = contract.get("item_sca_macro", {})
        item_names = contract.get("item_names", {})
        itens = [{
            "item_cod": r["item_cod"], "item": item_names.get(r["item_cod"], r["item"]), "sca_macro": item_macro.get(r["item_cod"], r["sca_macro"]), "disciplina": r["disciplina"],
            "situacao": r["situacao"], "supervisor": r["supervisor"], "gerente": r["gerente"],
            "previsto": item_date(r["previsto"], r["item_cod"], "previsto", contract), "replan": item_date(r["replan"], r["item_cod"], "replan", contract)
        } for r in rows]
        display_status = "Não iniciada" if status == "Não iniciado" else ("Concluída" if status == "Concluído" else status)
        configured_code = next((code for name, code in contract.get("area_codes", {}).items() if key(name) == area_key), str(number))
        display_exec = latest_date([r["replan"] or r["previsto"] for r in rows])
        display_prev = latest_date([r["previsto"] for r in rows])
        display_replan = latest_date([r["replan"] for r in rows])
        date_defaults = next((value for name, value in contract.get("area_dates", {}).items() if key(name) == area_key), {})
        if not (display_exec and re.fullmatch(r"\d{2}/\d{2}/\d{4}", display_exec)): display_exec = date_defaults.get("prazo_exec")
        if not (display_prev and re.fullmatch(r"\d{2}/\d{2}/\d{4}", display_prev)): display_prev = date_defaults.get("prazo_prev")
        if not (display_replan and re.fullmatch(r"\d{2}/\d{2}/\d{4}", display_replan)): display_replan = date_defaults.get("prazo_replan")
        areas.append({
            "area": area_names[area_key], "cod": configured_code, "macro": rows[0]["macro"], "ilha": rows[0]["ilha"], "processo": rows[0]["processo"],
            "fase": rows[0]["fase"], "itens": itens, "n_itens": len(rows), "concluidos": conc, "andamento": andamento,
            "nao_iniciados": nao, "progresso": progresso, "status": display_status,
            "prazo_exec": display_exec, "prazo_prev": display_prev,
            "fase_bucket": rows[0]["fase"], "projeto": photos["Projeto"], "campo": photos["Obra"],
            "pend_ppt": next((value for name, value in contract.get("area_pend_ppt", {}).items() if key(name) == area_key), [r["observacao"] for r in rows if r["observacao"] and r["situacao"] != "Concluído"]),
            "prazo_replan": display_replan,
        })
    result["areas"] = areas
    area_counts = Counter(a["status"] for a in areas)
    item_counts = Counter(r["situacao"] for r in raw)
    result["totais"] = {
        "total_areas": len(areas), "ilhas": len({key(r["ilha"]) for r in raw}), "total_scas": len(raw),
        "areas_andamento": area_counts["Em andamento"], "areas_nao_iniciadas": area_counts["Não iniciada"], "areas_concluidas": area_counts["Concluída"],
        "itens_concluidos": item_counts["Concluído"], "itens_andamento": item_counts["Em andamento"], "itens_nao_iniciados": item_counts["Não iniciado"],
    }
    replanned = sum(1 for r in raw if r["replan"])
    result["reprog"] = {"scas": replanned, "pct": round(100 * replanned / len(raw), 1), "total": replanned}

    def grouped_summary(field: str, output_name: str) -> list[dict[str, Any]]:
        buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        names: dict[str, str] = {}
        for r in raw:
            k = key(r[field]); names.setdefault(k, r[field]); buckets[k].append(r)
        out = []
        for k, rows in buckets.items():
            c, a, n, p, st = aggregate_status(rows)
            out.append({output_name: names[k], "n": len(rows), "conc": c, "andamento": a, "nao": n, "pend": a+n, "pendentes": a+n, "progresso": p, "pct_conc": round(100*c/len(rows), 1), "status": st, "prazo_exec": latest_date([r["replan"] or r["previsto"] for r in rows]), "areas": len({key(r["area"]) for r in rows})})
        return sorted(out, key=lambda x: (-x["n"], key(x[output_name])))
    result["macro"] = [{k: x[k] for k in ("macro", "n", "conc", "progresso", "pct_conc", "pendentes")} for x in grouped_summary("macro", "macro")]
    result["disciplina"] = [{k: x[k] for k in ("disciplina", "n", "conc", "andamento", "nao", "progresso")} for x in grouped_summary("disciplina", "disciplina")]
    reference = contract.get("sca_macro_reference", [])
    if reference and sum(x.get("n", 0) for x in reference) == len(raw):
        result["sca_macro"] = reference
    else:
        result["sca_macro"] = [{k: x[k] for k in ("macro", "n", "conc", "andamento", "nao", "pend", "progresso", "areas", "status", "prazo_exec")} for x in grouped_summary("sca_macro", "macro")]
    manager = grouped_summary("gerente", "gerente")
    total_pending = sum(x["pend"] for x in manager)
    result["gerente"] = [{"gerente": x["gerente"], "pend": x["pend"], "pct": round(100*x["pend"]/total_pending, 1) if total_pending else 0} for x in manager]
    phase_buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in raw: phase_buckets[r["fase"]].append(r)
    result["fase"] = {name: {"total": len(rows), "conc": aggregate_status(rows)[0], "pct": round(100*aggregate_status(rows)[0]/len(rows), 1)} for name, rows in phase_buckets.items()}
    return result


def empty_documents() -> dict[str, Any]:
    return {"source": None, "deadline": "Não informado", "columns": {"area": "Setor responsável", "manager": "Liderança responsável", "comments": "Observações"}, "filterLabels": {"area": "Todos os setores responsáveis", "manager": "Todas as lideranças responsáveis"}, "doneLabel": "Entregue", "rows": []}


def read_documents(folder: Path, contract: dict[str, Any]) -> dict[str, Any]:
    result = empty_documents()
    result.update(contract.get("documents", {}))
    path = first_workbook(folder)
    if not path:
        return result
    result["source"] = f"Checklist LO · {path.name}"
    ws = read_sheet(path)
    found = find_header(ws, {"Descrição", "Status"})
    if not found:
        return result
    header_row, headers = found
    desc_i = header_index(headers, "Descrição")
    area_i = header_index(headers, "Setor Responsavel", "Setor Responsável")
    manager_i = header_index(headers, "Liderança Responsável")
    status_i = header_index(headers, "Status")
    comments_i = header_index(headers, "Observações")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        description = clean_cell(value_at(row, desc_i))
        if not description:
            continue
        rows.append({"description": description, "area": clean_cell(value_at(row, area_i), "Não informado"), "manager": clean_cell(value_at(row, manager_i), "Não informado"), "status": clean_cell(value_at(row, status_i), "Pendente"), "comments": clean_cell(value_at(row, comments_i))})
    result["rows"] = rows
    return result


def main() -> int:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    all_images: list[dict[str, Any]] = []
    manifest = {"schemaVersion": 1, "units": {}}
    for unit_id, spec in UNITS.items():
        folder = SOURCE_ROOT / spec["folder"]
        images, image_map = build_images(unit_id, folder)
        all_images.extend(images)
        contract = CONTRACT.get(unit_id, {})
        operational = read_operational(first_workbook(folder), spec["label"], image_map, contract)
        if operational["areas"]:
            source_meta = dict(operational["meta"])
            operational["meta"] = dict(contract.get("meta", source_meta))
            operational["meta"]["atualizacao"] = source_meta["atualizacao"]
            operational["meta"]["responsavel"] = person_name(source_meta["responsavel"], contract.get("meta", {}).get("responsavel", "Não informado"))
        else:
            operational["meta"] = contract.get("meta", operational["meta"])
        documents = read_documents(folder / "3. Documentações LO", contract)
        payload = {"label": spec["label"], "operational": operational, "documents": documents}
        filename = f"{unit_id}.json"
        (OUTPUT_ROOT / filename).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest["units"][unit_id] = filename
    (OUTPUT_ROOT / "imagens.json").write_text(json.dumps({"schemaVersion": 1, "imagens": all_images}, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_ROOT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Gerados {len(UNITS)} arquivos de unidade e manifesto com {len(all_images)} imagens.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
